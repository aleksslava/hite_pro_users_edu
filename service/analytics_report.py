from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from db.models import Click, Session as SessionRow, User


USER_COLUMNS = [
    "id", "is_admin", "tg_user_id", "max_user_id", "username", "first_name", "last_name",
    "amo_contact_id", "amo_deal_id", "utm_campaign", "utm_medium", "utm_content",
    "utm_term", "utm_source", "yclid", "start_parameter", "phone_number", "created_at",
    "notification_stage", "current_state",
]

SESSION_COLUMNS = [
    "id", "user_id", "created_at", "last_activity_at", "open_until_at", "closed_at",
    "is_closed", "extension_window_minutes", "close_after_last_activity_minutes",
]

CLICK_COLUMNS = [
    "id", "session_id", "created_at", "dialog_window", "dialog_button", "weight",
]


def _naive(value):
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.astimezone(timezone.utc).replace(tzinfo=None)
    return value


def _write_rows(ws, columns: list[str], rows: list) -> None:
    ws.append(columns)
    for row in rows:
        ws.append([_naive(getattr(row, col)) for col in columns])


async def _fill_users(ws, db: AsyncSession) -> None:
    users = (await db.execute(select(User).order_by(User.id))).scalars().all()
    _write_rows(ws, USER_COLUMNS, users)


async def _fill_sessions(ws, db: AsyncSession) -> None:
    sessions = (await db.execute(select(SessionRow).order_by(SessionRow.id))).scalars().all()
    _write_rows(ws, SESSION_COLUMNS, sessions)


async def _fill_clicks(ws, db: AsyncSession) -> None:
    clicks = (await db.execute(select(Click).order_by(Click.id))).scalars().all()
    _write_rows(ws, CLICK_COLUMNS, clicks)


async def _fill_charts(ws, db: AsyncSession) -> None:
    reg_date = func.date(User.created_at)
    reg_rows = (
        await db.execute(
            select(reg_date.label("d"), func.count().label("c"))
            .where(User.created_at.is_not(None))
            .group_by(reg_date)
            .order_by(reg_date)
        )
    ).all()

    src_rows = (
        await db.execute(
            select(
                func.coalesce(func.nullif(User.utm_source, ""), "(none)").label("s"),
                func.count().label("c"),
            )
            .group_by("s")
            .order_by(func.count().desc())
        )
    ).all()

    click_rows = (
        await db.execute(
            select(
                func.coalesce(Click.dialog_window, "(none)").label("w"),
                func.count().label("c"),
            )
            .group_by(Click.dialog_window)
            .order_by(func.count().desc())
            .limit(10)
        )
    ).all()

    sess_date = func.date(SessionRow.created_at)
    duration_seconds = func.extract("epoch", SessionRow.closed_at - SessionRow.created_at)
    dur_rows = (
        await db.execute(
            select(sess_date.label("d"), func.avg(duration_seconds).label("avg_s"))
            .where(SessionRow.is_closed.is_(True))
            .group_by(sess_date)
            .order_by(sess_date)
        )
    ).all()

    ws["A1"] = "Регистрации по дням"
    ws["A2"] = "Дата"
    ws["B2"] = "Пользователей"
    for i, row in enumerate(reg_rows, start=3):
        ws.cell(row=i, column=1, value=str(row.d))
        ws.cell(row=i, column=2, value=int(row.c))

    if reg_rows:
        chart = LineChart()
        chart.title = "Регистрации пользователей по дням"
        chart.x_axis.title = "Дата"
        chart.y_axis.title = "Пользователей"
        data_ref = Reference(ws, min_col=2, min_row=2, max_row=2 + len(reg_rows))
        cat_ref = Reference(ws, min_col=1, min_row=3, max_row=2 + len(reg_rows))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cat_ref)
        ws.add_chart(chart, "D2")

    src_start_row = 3 + len(reg_rows) + 2
    ws.cell(row=src_start_row - 1, column=1, value="Источники трафика (utm_source)")
    ws.cell(row=src_start_row, column=1, value="Источник")
    ws.cell(row=src_start_row, column=2, value="Пользователей")
    for i, row in enumerate(src_rows, start=src_start_row + 1):
        ws.cell(row=i, column=1, value=str(row.s))
        ws.cell(row=i, column=2, value=int(row.c))

    if src_rows:
        chart = PieChart()
        chart.title = "Источники трафика"
        data_ref = Reference(ws, min_col=2, min_row=src_start_row, max_row=src_start_row + len(src_rows))
        cat_ref = Reference(ws, min_col=1, min_row=src_start_row + 1, max_row=src_start_row + len(src_rows))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cat_ref)
        ws.add_chart(chart, f"D{src_start_row}")

    click_start_row = src_start_row + len(src_rows) + 3
    ws.cell(row=click_start_row - 1, column=1, value="Топ окон диалога по кликам")
    ws.cell(row=click_start_row, column=1, value="Окно")
    ws.cell(row=click_start_row, column=2, value="Кликов")
    for i, row in enumerate(click_rows, start=click_start_row + 1):
        ws.cell(row=i, column=1, value=str(row.w))
        ws.cell(row=i, column=2, value=int(row.c))

    if click_rows:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Клики по окнам диалога (топ-10)"
        chart.x_axis.title = "Кликов"
        chart.y_axis.title = "Окно"
        data_ref = Reference(ws, min_col=2, min_row=click_start_row, max_row=click_start_row + len(click_rows))
        cat_ref = Reference(ws, min_col=1, min_row=click_start_row + 1, max_row=click_start_row + len(click_rows))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cat_ref)
        ws.add_chart(chart, f"D{click_start_row}")

    dur_start_row = click_start_row + len(click_rows) + 3
    ws.cell(row=dur_start_row - 1, column=1, value="Средняя длительность сессии по дням (мин)")
    ws.cell(row=dur_start_row, column=1, value="Дата")
    ws.cell(row=dur_start_row, column=2, value="Средняя длительность, мин")
    for i, row in enumerate(dur_rows, start=dur_start_row + 1):
        ws.cell(row=i, column=1, value=str(row.d))
        ws.cell(row=i, column=2, value=float(row.avg_s or 0) / 60.0)

    if dur_rows:
        chart = LineChart()
        chart.title = "Средняя длительность сессии по дням"
        chart.x_axis.title = "Дата"
        chart.y_axis.title = "Минуты"
        data_ref = Reference(ws, min_col=2, min_row=dur_start_row, max_row=dur_start_row + len(dur_rows))
        cat_ref = Reference(ws, min_col=1, min_row=dur_start_row + 1, max_row=dur_start_row + len(dur_rows))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cat_ref)
        ws.add_chart(chart, f"D{dur_start_row}")


async def build_analytics_xlsx(db: AsyncSession) -> bytes:
    wb = Workbook()
    users_ws = wb.active
    users_ws.title = "Пользователи"
    sessions_ws = wb.create_sheet("Сессии")
    clicks_ws = wb.create_sheet("Клики")
    charts_ws = wb.create_sheet("Инфографика")

    await _fill_users(users_ws, db)
    await _fill_sessions(sessions_ws, db)
    await _fill_clicks(clicks_ws, db)
    await _fill_charts(charts_ws, db)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
